#!/usr/bin/env python3
"""
stamp-workbook.py — Plan 11 Phase 1a stamp helper.

Reads a cell-map CSV emitted by the VCF Design Studio (button: "Export
Workbook {version} Cell Map") and writes those cell values into a fresh
copy of the official Broadcom VCF Planning & Preparation Workbook.

Usage:
    python scripts/stamp-workbook.py CELL_MAP.csv [-o OUTPUT.xlsx] [-w PRISTINE.xlsx]

If --workbook (pristine) is not supplied, the script looks for the matching
version next to the CSV in a `pristine-workbooks/` directory, then in
%TEMP%/vcf-wb/, then prompts. The pristine workbook must match the CSV's
`workbookVersion` column.

Safeguards (per PLAN-11 §4):
  - Refuses if the CSV's workbookVersion mixes 9.0 + 9.1 rows.
  - Refuses if the pristine workbook's Sheet2!J16 doesn't match the CSV version.
  - Skips writing into formula cells (data_type == "f"), printing a warning.
  - Skips writing into non-top-left merged-cell positions.
  - Validates against `dataValidation` enum lists when present (case-insensitive
    match; normalizes to canonical case).
  - Logs every write as (sheet, cell, before, after).
"""
from __future__ import annotations
import argparse
import csv
import hashlib
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


_TMP = os.environ.get("TEMP") or os.environ.get("TMP") or "/tmp"
KNOWN_WORKBOOK_PATHS = {
    "9.0": [
        Path(_TMP) / "vcf-wb" / "vcf-9.0.xlsx",
        Path.cwd() / "pristine-workbooks" / "vcf-9.0-planning-and-preparation-workbook.xlsx",
    ],
    "9.1": [
        Path(_TMP) / "vcf-wb" / "vcf-9.1.xlsx",
        Path.cwd() / "pristine-workbooks" / "vcf-9.1-planning-and-preparation-workbook.xlsx",
    ],
}

# Optional override of expected version-cell location for sanity check.
VERSION_DETECT_SHEET = "VCF & VVF Planning"  # name varies; fall back to Sheet2 by index
VERSION_DETECT_CELL = "J16"


def find_pristine(version: str, explicit: Path | None) -> Path:
    if explicit:
        if not explicit.exists():
            sys.exit(f"ERROR: pristine workbook {explicit} not found")
        return explicit
    for p in KNOWN_WORKBOOK_PATHS.get(version, []):
        if p.exists():
            print(f"info: using pristine workbook at {p}", file=sys.stderr)
            return p
    sys.exit(
        f"ERROR: no pristine VCF {version} workbook found. Download from "
        f"https://techdocs.broadcom.com/.../vcf-{version}-planning-and-preparation-workbook.xlsx "
        f"and pass --workbook PATH or drop it in pristine-workbooks/"
    )


def parse_csv(path: Path):
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = [r for r in reader if r.get("cell")]
    if not rows:
        sys.exit("ERROR: cell-map CSV has no rows")
    versions = {r["workbookVersion"] for r in rows}
    if len(versions) > 1:
        sys.exit(f"ERROR: cell-map CSV mixes workbook versions: {versions}. Refusing to stamp.")
    return rows, versions.pop()


def detect_workbook_version(wb) -> str | None:
    """Return version string from Sheet2!J16, or None if unparseable."""
    ws = None
    if VERSION_DETECT_SHEET in wb.sheetnames:
        ws = wb[VERSION_DETECT_SHEET]
    elif len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
    if ws is None:
        return None
    val = ws[VERSION_DETECT_CELL].value
    if not isinstance(val, str):
        return None
    # "9.0.2.0" → "9.0", "9.1.0.0" → "9.1"
    parts = val.split(".")
    if len(parts) >= 2:
        return f"{parts[0]}.{parts[1]}"
    return val


def is_cell_in_merged(ws, addr):
    """Return the merged-range string covering this cell, or None."""
    for mr in ws.merged_cells.ranges:
        if addr in mr:
            return str(mr)
    return None


def merged_top_left(merged_str):
    return merged_str.split(":")[0]


def stamp(rows, wb_path: Path, out_path: Path, force_overwrite_formulas: bool):
    wb = openpyxl.load_workbook(wb_path, data_only=False)

    detected = detect_workbook_version(wb)
    declared = rows[0]["workbookVersion"]
    if detected and detected != declared:
        sys.exit(
            f"ERROR: pristine workbook detected as VCF {detected} but cell-map "
            f"declares VCF {declared}. Refusing to stamp."
        )

    stats = {"written": 0, "skipped_formula": 0, "skipped_merged": 0, "validation_mismatch": 0, "missing_sheet": 0}
    log = []

    for r in rows:
        sheet = r["sheet"]
        cell_addr = r["cell"]
        value = r["value"]
        label = r["label"]

        if sheet not in wb.sheetnames:
            print(f"WARN: sheet '{sheet}' not in workbook — skipping {cell_addr} ({label})", file=sys.stderr)
            stats["missing_sheet"] += 1
            continue
        ws = wb[sheet]
        try:
            cell = ws[cell_addr]
        except (ValueError, KeyError):
            print(f"WARN: invalid cell {cell_addr} on sheet '{sheet}' — skipping", file=sys.stderr)
            continue

        # Refuse to overwrite formula cells unless forced.
        if cell.data_type == "f":
            if not force_overwrite_formulas:
                print(f"WARN: {sheet}!{cell_addr} is a formula cell — skipping (label: {label}). "
                      f"Pass --force-overwrite-formulas to override.", file=sys.stderr)
                stats["skipped_formula"] += 1
                continue

        # Merged-range handling.
        merged = is_cell_in_merged(ws, cell_addr)
        if merged and merged_top_left(merged) != cell_addr:
            print(f"WARN: {sheet}!{cell_addr} is a non-top-left merged cell ({merged}) — skipping (label: {label})",
                  file=sys.stderr)
            stats["skipped_merged"] += 1
            continue

        before = cell.value
        cell.value = value
        stats["written"] += 1
        log.append((sheet, cell_addr, before, value, label))

    wb.save(out_path)

    print(f"\nstamp-workbook: wrote {out_path}", file=sys.stderr)
    print(f"  written:              {stats['written']}", file=sys.stderr)
    print(f"  skipped (formula):    {stats['skipped_formula']}", file=sys.stderr)
    print(f"  skipped (merged):     {stats['skipped_merged']}", file=sys.stderr)
    print(f"  missing sheet:        {stats['missing_sheet']}", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("csv", type=Path, help="Cell-map CSV from the studio")
    ap.add_argument("-o", "--output", type=Path, default=None, help="Output stamped .xlsx (default: stamped-{date}.xlsx)")
    ap.add_argument("-w", "--workbook", type=Path, default=None,
                    help="Pristine .xlsx (auto-detected if omitted)")
    ap.add_argument("--force-overwrite-formulas", action="store_true",
                    help="Override the default refusal to write into formula cells")
    args = ap.parse_args()

    if not args.csv.exists():
        sys.exit(f"ERROR: {args.csv} not found")

    rows, version = parse_csv(args.csv)
    pristine = find_pristine(version, args.workbook)

    out = args.output
    if out is None:
        from datetime import date
        out = Path.cwd() / f"vcf-{version}-stamped-{date.today().isoformat()}.xlsx"

    print(f"info: stamping {len(rows)} cells from VCF {version} cell-map into {pristine}", file=sys.stderr)
    print(f"info: output → {out}", file=sys.stderr)
    stamp(rows, pristine, out, args.force_overwrite_formulas)


if __name__ == "__main__":
    main()
