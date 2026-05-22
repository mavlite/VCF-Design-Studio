#!/usr/bin/env python3
"""
extract-workbook-cell-meta.py — extract cell metadata from VCF Planning & Preparation Workbooks.

Produces test-fixtures/workbook/workbook-cell-meta-{version}.json with every
labeled user-input cell on the studio-relevant sheets, capturing:
  - sheet name
  - cell address
  - companion label cell + text
  - data type (s / n / f / etc per openpyxl)
  - sample value
  - data validation enum (if cell is in a list-validation range)
  - merged-range membership (if any)

Plus Sheet2!J16 (the version cell used for import-side version detection) and
the vCenter Medium default storage reference cell (C52 in 9.0 / C55 in 9.1).

Run:
    python scripts/extract-workbook-cell-meta.py
"""
from __future__ import annotations
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

import os
_TMP = os.environ.get("TEMP") or os.environ.get("TMP") or "/tmp"
WORKBOOKS = [
    ("9.0", Path(_TMP) / "vcf-wb" / "vcf-9.0.xlsx"),
    ("9.1", Path(_TMP) / "vcf-wb" / "vcf-9.1.xlsx"),
]

# Sheets the studio plans to read/write, by canonical name. The same sheet
# may live at different sheet indices across workbook versions; we look up
# by name so the script is version-stable.
#
# Each entry maps the canonical sheet name to:
#   - label_col: the column letter holding the label text
#   - value_col: the column letter the stamp script targets (user-input cell)
#   - row_range: (min_row, max_row) — empirical bounds of populated rows
SHEET_TARGETS = {
    "Deploy Management Domain": {
        "label_col": "J",
        "value_col": "L",
        "row_range": (1, 450),
    },
    "Configure Management Domain": {
        "label_col": "B",
        "value_col": "D",
        "row_range": (1, 600),
    },
    "Deploy Workload Domain": {
        # This sheet uses B = label, C = sample, D = user-input column
        # (same convention as Configure sheets).
        # The top "Select Option / Feature / Final Result / Information"
        # block at rows 9-18 is a control-knob area we deliberately skip
        # (its B column is pre-populated, not a user-input cell).
        "label_col": "B",
        "value_col": "D",
        "row_range": (1, 900),
    },
    "Configure Workload Domain": {
        "label_col": "B",
        "value_col": "D",
        "row_range": (1, 300),
    },
    "Deploy Cluster": {
        # Same B = label / D = value convention.
        "label_col": "B",
        "value_col": "D",
        "row_range": (1, 500),
    },
}

# Static / fixed-address cells the studio reads but does not write
# (version detection + cross-checks).
READ_ONLY_TARGETS = [
    {
        "sheet_index": 2,   # Sheet2
        "fallback_sheet_name_substring": None,  # Sheet2 has no canonical name
        "cell": "J16",
        "purpose": "version-detect",
        "note": "Carries the literal workbook version string (e.g. '9.0.2.0' or '9.1.0.0').",
    },
    {
        "sheet_name_substring": "Static Reference Tables",
        "cell_by_version": {"9.0": "C52", "9.1": "C55"},
        "purpose": "version-detect-corroborate",
        "note": "vCenter Medium default storage reference value. 908 in 9.0, 858 in 9.1.",
    },
]

def cell_at(ws, col_letter, row):
    return ws[f"{col_letter}{row}"]

def normalize_text(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return value

def find_validation_for_cell(ws, cell_addr):
    """Return the data-validation enum (list) covering this cell, or None."""
    if ws.data_validations is None:
        return None
    row, col = openpyxl.utils.cell.coordinate_to_tuple(cell_addr)
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        # dv.sqref is a MultiCellRange — iterate its CellRange members and
        # check bounds directly.
        hit = False
        for rng in dv.sqref.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                hit = True
                break
        if not hit:
            continue
        # Parse formula1 — could be "'Lists'!$A$1:$A$5" or a literal
        # "\"Small,Medium,Large,X-Large\"".
        f = dv.formula1 or ""
        if f.startswith('"') and f.endswith('"'):
            return [s.strip() for s in f[1:-1].split(",") if s.strip()]
        m = re.match(r"^(?:'?([^'!]+)'?!)?(\$?[A-Z]+\$?\d+):(\$?[A-Z]+\$?\d+)$", f)
        if m:
            sheet_name = m.group(1)
            start = m.group(2).replace("$", "")
            end = m.group(3).replace("$", "")
            try:
                target_ws = ws.parent[sheet_name] if sheet_name else ws
                values = []
                for r in target_ws[f"{start}:{end}"]:
                    for c in r:
                        if c.value is not None:
                            values.append(str(c.value))
                return values or None
            except (KeyError, ValueError):
                return None
        return None
    return None

def find_merged_range(ws, cell_addr):
    """Return the merged-range string covering this cell, or None."""
    for mr in ws.merged_cells.ranges:
        if cell_addr in mr:
            return str(mr)
    return None

def cell_data_type(cell):
    """Return openpyxl data_type: 's', 'n', 'b', 'f', 'd', 'e', 'inlineStr', or 'n' for None."""
    return cell.data_type

def find_sheet_by_name_substring(wb, substr):
    for name in wb.sheetnames:
        if substr.lower() in name.lower():
            return wb[name]
    return None

def extract_workbook(version, xlsx_path):
    print(f"[{version}] Loading {xlsx_path}…", file=sys.stderr)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    entries = []

    # 1) Iterate the studio-relevant sheets by name.
    for sheet_name, cfg in SHEET_TARGETS.items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else find_sheet_by_name_substring(wb, sheet_name)
        if ws is None:
            print(f"[{version}]   WARN: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        actual_name = ws.title
        label_col = cfg["label_col"]
        value_col = cfg["value_col"]
        rmin, rmax = cfg["row_range"]
        print(f"[{version}]   {actual_name}: scanning {label_col}{rmin}–{label_col}{rmax}", file=sys.stderr)
        for r in range(rmin, rmax + 1):
            label_cell = cell_at(ws, label_col, r)
            label = normalize_text(label_cell.value)
            if not isinstance(label, str):
                continue
            # Skip pure-header rows (no companion value cell intention) by
            # requiring the value cell to be writable (not in a merged
            # non-top-left position). We still emit if the value cell is
            # empty — many user-input cells are blank in the pristine
            # workbook by design.
            value_cell = cell_at(ws, value_col, r)
            value_addr = f"{value_col}{r}"
            merged = find_merged_range(ws, value_addr)
            # Skip when the value cell is part of a merged range and the
            # top-left of that range is NOT our value_addr (i.e., we're a
            # non-top-left merged cell — writing here raises in openpyxl).
            if merged:
                top_left = merged.split(":")[0]
                if top_left != value_addr:
                    continue
            entries.append({
                "sheet": actual_name,
                "cell": value_addr,
                "labelCell": f"{label_col}{r}",
                "labelText": label,
                "dataType": cell_data_type(value_cell),
                "sampleValue": value_cell.value if not isinstance(value_cell.value, (int, float)) else value_cell.value,
                "sampleCell": f"{get_column_letter(column_index_from_string(label_col) + 1)}{r}",
                "sampleCellValue": cell_at(ws, get_column_letter(column_index_from_string(label_col) + 1), r).value,
                "sampleCellDataType": cell_at(ws, get_column_letter(column_index_from_string(label_col) + 1), r).data_type,
                "dataValidation": find_validation_for_cell(ws, value_addr),
                "mergedRange": merged,
            })

    # 2) Read-only targets (version cell + storage reference).
    for ro in READ_ONLY_TARGETS:
        ws = None
        if "sheet_index" in ro:
            # Sheet2 etc — by position
            try:
                ws = wb.worksheets[ro["sheet_index"] - 1]
            except IndexError:
                pass
        elif "sheet_name_substring" in ro:
            ws = find_sheet_by_name_substring(wb, ro["sheet_name_substring"])
        if ws is None:
            print(f"[{version}]   WARN: read-only target {ro} sheet not found", file=sys.stderr)
            continue
        addr = ro.get("cell") or ro["cell_by_version"][version]
        c = ws[addr]
        entries.append({
            "sheet": ws.title,
            "cell": addr,
            "purpose": ro["purpose"],
            "note": ro["note"],
            "dataType": cell_data_type(c),
            "sampleValue": c.value,
        })

    return {
        "workbookVersion": version,
        "extractedFrom": str(xlsx_path),
        "sha256": hashlib.sha256(xlsx_path.read_bytes()).hexdigest(),
        "sheetNames": list(wb.sheetnames),
        "entries": entries,
    }

def main():
    out_dir = Path(__file__).parent.parent / "test-fixtures" / "workbook"
    out_dir.mkdir(parents=True, exist_ok=True)
    for version, xlsx in WORKBOOKS:
        if not xlsx.exists():
            print(f"ERROR: missing {xlsx}", file=sys.stderr)
            sys.exit(1)
        meta = extract_workbook(version, xlsx)
        out = out_dir / f"workbook-cell-meta-{version}.json"
        with out.open("w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False, default=str)
        print(f"[{version}] wrote {out} ({len(meta['entries'])} entries)", file=sys.stderr)

if __name__ == "__main__":
    main()
